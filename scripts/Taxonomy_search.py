# #!/usr/bin/python3
############################# INTRODUCTION #############################
# Author: Guillaume Le Berre
# GitHub: https://github.com/GuillaumeLeBerreBIT
# 
# Script that opens the excel file "Species_Macrobenthos.xlsx" and iterates over the rows to download the sequences available for the different species. 
# The NCBI search will be a taxonomy search on Nucleotide database of NCBI, each species will ahve a database with all possible sequences available. 
#
# https://www.ncbi.nlm.nih.gov/books/NBK25499/#chapter4.EFetch
#
# Path to the Excel datafile: "C:\Users\gleberre\OneDrive - ILVO\Documenten\05_BAR-kustvloot\02_experiments\01_genome_skimming\01_inputs"
###################################################################
# 
############################# MODULES #############################
# Importing all necassary librarys to execute the script. 

from Bio import Entrez
from openpyxl import load_workbook # pip install openpyxl
import argparse

############################# COMMAND LINE #############################
parser = argparse.ArgumentParser(description='Download the sequence information of the Taxonomic ID')
parser.add_argument('inputFile', type=str, 
                    help='Give the (path to and) name of the excel file containing the taxonomic data.')
parser.add_argument('outputFolder', type=str, 
                    help='Give the Fasta file containing the filtered reads on a certain threshold.')

parser.add_argument('-i', '--TaxID', type=str, default = "B", required = False, 
                    help ='The column e.g. A, B, C, ... where the taxonomic ID is present')
parser.add_argument('-s', '--Species_name', type=str, default = "A", required = False, 
                    help ='The column e.g. A, B, C, ... where the Species name is present')
parser.add_argument('-a', '--Api_Key', type=str, default = "", required = False, 
                    help ='Give your Api_Key to get more querry requests.')
parser.add_argument('-e', '--email', type=str, default = "", required = False, 
                    help ='Give own email adress which the Api_Key is used from.')
parser.add_argument('-q', '--query', type=str, default = "All", required = False, 
                    help ='All: downloads all nucleotide taxa that can be found under the TaxID of that species. \
                        Complete: downloads the complete ribosomal DNA region 18S-ITS1-5.8S-ITS2-28S.\
                        Ribosomal: downloads any nucleotide sequences that have either 18S, 5.8S or 28S region present under the nucleotide taxonomy ID.')
args = parser.parse_args()

############################# MAIN SCRIPT #############################

# Add at the start to not continously call API key
# Using my perosnal API Key for up to 10 querry requests per second
# "b8a3b335a894e549dc83fa4b4b500cd03608"
Entrez.api_key = args.Api_Key

# Use the optional email parameter so the NCBI can contact you if there is a problem. You can either explicitly
# set this as a parameter with each call to Entrez
# "uniroguillaumelb@gmail.com"
Entrez.email = args.email

# You can also use forward slashes instead of backslashes in the file path, which is a platform-independent approach and works in Python on Windows
# "C:/Users/gleberre/OneDrive - ILVO/Documenten/05_BAR-kustvloot/02_experiments/01_genome_skimming/01_inputs/Species_Macrobenthos.xlsx"
wb = load_workbook(filename = args.inputFile)
# Gets the first worksheet, which here containts only one worksheet
first_sheet = wb.sheetnames[0]
# Which is usefull to do if multiple sheets present, now one sheet does not really matter.
worksheet = wb[first_sheet]
# The species name is in the first column & The taxonomic ID is in the 2nd column of the excel file. 
# "A", "B"
Species_col, Tax_ID_col = args.Species_name, args.TaxID

# Here you iterate over the rows in the specific column, start from row 2 since row 1 is a header line
for row in range(2,worksheet.max_row + 1):  
    # Create the cell name to call the value from e.g. B4, B16, ...
    cell_name_taxid = "{}{}".format(Tax_ID_col, row)      
    Tax_ID = worksheet[cell_name_taxid].value # the value of the specific cell
    Stripped_TaxID = str(Tax_ID).strip()
    
    # The search query that downloads all nucleotide taxa that can be found under the TaxID of that species
    if args.query == "All" or\
        args.query == "all":
        search_query = f"txid{Stripped_TaxID}[Orgn]"
    # The search query that downloads the complete ribosomal DNA region 18S-ITS1-5.8S-ITS2-28S.
    elif args.query == "Complete" or\
         args.query == "complete": 
        search_query = f"txid{Stripped_TaxID}[Orgn] AND 18s ribosomal rna[All Fields] AND gene[All Fields] AND internal[All Fields] AND transcribed[All Fields]\
                        AND spacer[All Fields] AND 1[All Fields] AND 5 8s ribosomal rna[All Fields] AND gene[All Fields] AND internal[All Fields] AND transcribed[All Fields] \
                        AND spacer[All Fields] AND 2[All Fields] AND 28s ribosomal rna[All Fields] AND gene[All Fields] AND complete[All Fields] AND sequence[All Fields]"
    # The search query that download any nucleotide sequences that have either 18S, 5.8S or 28S region present under the nucleotide taxonomy ID. 
    elif args.query == "Ribosomal" or\
         args.query == "ribosomal": 
        search_query = f"txid{Stripped_TaxID}[Orgn] AND (ribosomal RNA[All Fields] AND 5.8S[All Fields] OR 18S[All Fields] OR 28S[All Fields]) NOT Chromosome"
    
    
    # Using the esearch to search databases for the wanted terms, being taxanomic ID "txid{}[Orgn]"
    Tax_ID_search = Entrez.esearch(db='nucleotide', term = search_query)
    
    # Using the Biopython parser to abtain a Python object instead of XML. -- > A dictionary which has multiple keys, want "IdList"
    Record_Tax_ID = Entrez.read(Tax_ID_search)
    
    # If there are no query's then go again to the next TaxID.
    if len(Record_Tax_ID["IdList"]) == 0:
        continue

    # Returns all the nucleotide IDs from taxonomic ID
    Record_Tax_ID["IdList"]

    # Can retrieve the a fullr ecord from Entrez using the IDs gathered from the Taxonomic search in Nucleotide database
    # rettype = FASTA -- > To get the FASTA sequentie
    # retmode = text -- > To get it in a text format file 
    Record = Entrez.efetch(db="nucleotide", 
                        id= Record_Tax_ID["IdList"], 
                        rettype= 'fasta', 
                        retmode = "text")
    # This will print the nucleotide sequence for all the taxonomic IDs
    #print(Record.read())

    Tax_ID_search.close()
    
    # Use the cell to call the species name as well
    cell_name_species = "{}{}".format(Species_col, row)      
    Species_name = worksheet[cell_name_species].value # the value of the specific cell
    # Replace the space in the string by "_"
    Species_name_replaced = Species_name.replace(" ","_")
    # Open an empty file to write to 
    with open(f"{args.outputFolder}/{Species_name_replaced}_DB.fasta", "w") as file_to_write: 
        # Write the sequences to a FASTA file
        file_to_write.write(Record.read())
    # Close the file
    file_to_write.close()